# -*-coding:utf-8 -*-
# @Author : Zhigang

import os
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font,colors

#openpyxl==2.4.5版本
class ParseExcel(object):

    def __init__(self,filePath):
        try:
            if os.path.abspath(filePath):
                self.filepath=filePath
            self.workbook=load_workbook(self.filepath)
            self.sheet=self.workbook.active
        except FileNotFoundError as e:
            #return None #不能返回字符串
            print ("文件未找到，必须为绝对路径")

    def get_sheet_by_index(self,sheetIndex):
        "通过sheet索引获取sheet对象"
        try:
            self.sheet=self.workbook[self.workbook.sheetnames[sheetIndex]]
        except IndexError as e:
            raise (e)
        else:
            return self.sheet

    def get_sheet_by_name(self,sheetName):
        "通过sheet名称获取sheet对象"
        try:
            self.sheet=self.workbook[sheetName]
        except KeyError as e:
            raise (e)
        else:
            return self.sheet

    def setSheetByIndex(self,sheetIndex):
        "设置当前要操作的sheet对象,使用index来获取相应的sheet"
        self.sheet=self.get_sheet_by_index(sheetIndex)

    def setSheetByName(self,sheetName):
        "设置当前要操作的sheet对象,使用sheetName来获取相应的sheet"
        # if sheetName in self.workbook.sheetnames:
        #     self.sheet=self.workbook[sheetName]
        self.sheet=self.get_sheet_by_name(sheetName)

    def get_default_name(self):
        "获取当前默认sheet的名字"
        return self.sheet.title

    def get_max_row_no(self):
        "获取默认sheet中最大的行数"
        return self.sheet.max_row

    def get_max_col_no(self):
        "获取默认sheet中最大的列数"
        return self.sheet.max_column

    def get_min_row_no(self):
        "获取默认sheet中起始行号,从1开始"
        return self.sheet.min_row

    def get_min_col_no(self):
        "获取默认sheet中起始列号,从1开始"
        return self.sheet.min_column

    def get_all_rows(self):
        "获取默认sheet的所有行对象,三种方式"
        # return list(self.sheet.rows)
        # for row in self.sheet.rows:
        #     print (row)
        rows=[]
        for row in self.sheet.iter_rows():
            rows.append(row)
        return rows

    def get_all_cols(self):
        "获取默认sheet中的所有列对象,三种方式"
        # return list(self.sheet.columns)
        # for column in self.sheet.columns:
        #     print (column)
        cols=[]
        for col in self.sheet.iter_cols():
            cols.append(col)
        return cols

    def get_single_row(self,row_no):
        "获取默认sheet中的某一行,起始行为1"
        # if not isinstance(row_no,int):
        #     return "参数须为数字"
        # return self.sheet[row_no]
        # for cell in self.sheet[row_no]:
        #     print (cell.value)
        # return list(self.sheet.rows)[row_no-1] #行号减1等于索引位置
        return self.get_all_rows()[row_no-1]

    def get_single_col(self,col_no):
        # "获取默认sheet中的某一列，起始列为'A'"
        # if not isinstance(col_no,str):
        #     return "参数须为'A','B'样式"
        # return self.sheet[col_no]
        # for col in self.sheet[col_no]:
        #     print (col.value)
        # return list(self.sheet.columns)[col_no-1]  #列号减1等于索引位置
        return self.get_all_cols()[col_no-1]

    def get_cell(self,row_no,col_no):
        "从默认sheet中，通过行号和列号获取指定的单元格"
        return self.sheet.cell(row_no,col_no)

    def get_cell_content(self,row_no,col_no):
        "从默认sheet中，通过行号和列号获取指定单元格中的内容"
        return self.sheet.cell(row_no,col_no).value

    def write_cell_content(self,row_no,col_no,content,color=None):
        "从默认sheet中，通过行号和列号向指定单元格中写入指定内容,可指定颜色;写入文件中，文件须是关闭状态"
        # self.sheet.cell(row_no,col_no,value=content)
        cell=self.get_cell(row_no,col_no)
        cell.value=content
        cell.font=Font(color=color)
        self.save_excel_file()

    def write_cell_current_time(self,row_no,col_no):
        "从默认sheet中，通过行号和列号向指定单元格中写入当前日期"
        "一种方式"
        current_time=datetime.datetime.now()
        self.sheet.cell(row_no, col_no, value=current_time)
        "二种方式"
        # cell=self.get_cell(row_no,col_no)
        # cell.value=current_time
        "三种方式"
        # import time
        # current_time=time.strftime("%Y-%m-%d %H:%M:%S",time.localtime())
        # self.sheet.cell(row_no,col_no,value=current_time)
        self.save_excel_file()

    def save_excel_file(self):
        "保存文件"
        self.workbook.save(self.filepath)

if __name__=="__main__":
    f=ParseExcel("C:\\Users\\zhigang\\Desktop\\test.xlsx")
    #f=ParseExcel("test.xlsx")
    # print (f.get_sheet_by_index(1))
    # print (f.get_sheet_by_index(3))
    # print (f.get_sheet_by_name("test"))
    # print (f.get_sheet_by_name("dasf"))
    # f.setSheetByIndex(2)
    # print (f.get_default_name())
    # f.setSheetByName("tester")
    # print (f.get_default_name())
    # print (f.get_max_row_no())
    # print (f.get_max_col_no())
    # print (f.get_min_row_no())
    # print (f.get_min_col_no())
    # print (f.get_all_rows())
    # print (f.get_all_cols())
    # print (f.get_single_row(2))
    # print (f.get_single_row("3"))
    # print (f.get_single_col("B"))
    # print (f.get_single_col(3))
    # print (f.get_cell(2,3))
    # print (f.get_cell_content(3,4))
    # f.write_cell_content(8, 8, "Fail", color=colors.RED)
    # f.write_cell_content(9, 9, "True", color=colors.GREEN)
    # f.write_cell_content(1,1,"新年快乐")
    # f.write_cell_current_time(5,5)
    # f.save_excel_file()